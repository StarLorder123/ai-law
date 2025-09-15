import os
import json
import time
import pickle
from datetime import datetime
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Dict, Any, List
import re


class LawDataFetcher:
    def __init__(self, username=None, password=None):
        self.base_url = "https://bigdata.lawyee.net"
        self.login_url = "https://bigdata.lawyee.net/login"
        self.api_url = "https://bigdata.lawyee.net/dsjwjfx/ajts/getListByDataSource"

        self.username = "liuwei"
        self.password = "Runningwei@6"

        self.session = requests.Session()
        self.setup_logging()

        self.page_size = 10

        # Chrome配置
        self.chrome_options = webdriver.ChromeOptions()
        # self.chrome_options.add_argument('--headless')  # 无头模式，取消注释即可启用
        self.chrome_options.add_argument('--no-sandbox')
        self.chrome_options.add_argument('--disable-dev-shm-usage')

        # cookie文件路径
        self.cookie_file = 'law_cookies.pkl'
        self.session_file = 'law_session.pkl'

        # 添加结果文件路径
        self.result_file = 'law_results.txt'

    def setup_logging(self):
        """设置日志配置"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('law_data_fetch.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )

    def save_cookies_and_session(self, driver):
        """保存cookies和session"""
        try:
            # 保存cookies
            cookies = driver.get_cookies()
            with open(self.cookie_file, 'wb') as f:
                pickle.dump(cookies, f)

            # 保存session
            with open(self.session_file, 'wb') as f:
                pickle.dump(self.session.cookies, f)

            logging.info("成功保存cookies和session")
            return True
        except Exception as e:
            logging.error(f"保存cookies和session失败: {str(e)}")
            return False

    def load_cookies_and_session(self):
        """加载所有会话信息"""
        try:
            if os.path.exists(self.cookie_file) and os.path.exists(self.session_file):
                # 加载认证信息
                with open(self.cookie_file, 'rb') as f:
                    self.auth_info = pickle.load(f)

                # 加载session
                with open(self.session_file, 'rb') as f:
                    self.session.cookies.update(pickle.load(f))

                # 验证信息完整性
                required_keys = ['web_cookie', 'token', 'userInfo']
                if not all(key in self.auth_info for key in required_keys):
                    logging.error("会话信息不完整")
                    return False

                logging.info("成功加载所有会话信息")
                return True
            return False
        except Exception as e:
            logging.error(f"加载会话信息失败: {str(e)}")
            return False

    def get_web_cookie(self, driver):
        """获取web cookie值"""
        try:
            web_cookie = None
            cookies = driver.get_cookies()
            for cookie in cookies:
                if cookie['name'] == 'web':
                    web_cookie = cookie['value']
                    break
            return web_cookie
        except Exception as e:
            logging.error(f"获取web cookie失败: {str(e)}")
            return None

    def login(self):
        """使用Selenium进行登录"""
        driver = None
        try:
            driver = webdriver.Chrome(options=self.chrome_options)
            driver.get(self.login_url)

            driver.find_element(By.CLASS_NAME, "login-btn").click()

            # 等待登录表单加载
            wait = WebDriverWait(driver, 10)
            username_input = wait.until(EC.presence_of_element_located((By.NAME, "username")))
            password_input = driver.find_element(By.NAME, "password")

            # 输入账号密码
            username_input.send_keys(self.username)
            password_input.send_keys(self.password)

            time.sleep(5)

            # 点击登录按钮
            login_button = driver.find_element(By.ID, "btnSubmit")
            login_button.click()

            # 等待登录成功并跳转
            time.sleep(5)

            # 获取所有会话信息
            cookies = driver.get_cookies()
            local_storage = driver.execute_script("return window.localStorage;")
            session_storage = driver.execute_script("return window.sessionStorage;")

            # 获取web cookie和其他关键cookie
            web_cookie = self.get_web_cookie(driver)
            if not web_cookie:
                logging.error("未能获取web cookie")
                return False

            # 获取token和其他认证信息
            token = driver.execute_script("return localStorage.getItem('token');")
            user_info = driver.execute_script("return localStorage.getItem('userInfo');")

            # 保存完整的认证信息
            self.auth_info = {
                'cookies': cookies,
                'web_cookie': web_cookie,
                'localStorage': local_storage,
                'sessionStorage': session_storage,
                'token': token,
                'userInfo': user_info,
                'headers': {
                    'Authorization': token,
                    'Cookie': f'web={web_cookie}'
                }
            }

            # 更新session的cookies
            self.session.cookies.clear()
            for cookie in cookies:
                self.session.cookies.set(cookie['name'], cookie['value'])

            # 保存到文件
            with open(self.cookie_file, 'wb') as f:
                pickle.dump(self.auth_info, f)

            # 保存session状态
            with open(self.session_file, 'wb') as f:
                pickle.dump(self.session.cookies, f)

            logging.info("登录成功并保存所有会话信息")
            return True

        except Exception as e:
            logging.error(f"登录过程出错: {str(e)}")
            return False
        finally:
            if driver:
                driver.quit()

    def check_login_status(self):
        """检查登录状态"""
        try:
            response = self.session.get(f"{self.base_url}/homePage")
            return response.status_code == 200 and "login" not in response.url
        except:
            return False

    def fetch_detail(self, row_id):
        """获取法规详细信息"""
        try:
            detail_url = "https://bigdata.lawyee.net/dsjwjfx/qw/ybxml"
            
            headers = {
                'Accept': 'application/json, text/javascript, */*; q=0.01',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'Connection': 'keep-alive',
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                'Cookie': f'web={self.auth_info["web_cookie"]}',
                'Host': 'bigdata.lawyee.net',
                'Origin': 'https://bigdata.lawyee.net',
                'Referer': f'https://bigdata.lawyee.net/dsjwjfx/qw/pop-qw?rowKey={row_id}&ws_dl=flfg&type=1',
                'sec-ch-ua': '"Not(A:Brand";v="99", "Google Chrome";v="133", "Chromium";v="133"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'Sec-Fetch-Dest': 'empty',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Site': 'same-origin',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36',
                'X-Requested-With': 'XMLHttpRequest'
            }
            
            payload = {
                'rowkey': row_id,
                'dl': 'flfg',
                'highlightWord': '',
                'type': '1'
            }
            
            response = self.session.post(
                detail_url,
                headers=headers,
                data=payload,
                timeout=30
            )
            
            if response.status_code == 200:
                print(f"获取到ID为 {row_id} 的详细信息：")
                print(response.text)
                print("-" * 80)  # 分隔线
                data = json.loads(response.text)
                return data
            else:
                logging.error(f"获取详细信息失败: {response.status_code}")
                return None
            
        except Exception as e:
            logging.error(f"获取详细信息时出错: {str(e)}")
            return None

    def fetch_data(self, page=1, page_size=20):
        """获取法律数据"""
        try:
            if not hasattr(self, 'auth_info') or not self.auth_info.get('web_cookie'):
                if not self.load_cookies_and_session():
                    logging.error("缺少认证信息")
                    return None

            headers = {
                'Accept': '*/*',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'Connection': 'keep-alive',
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                'Cookie': f'web={self.auth_info["web_cookie"]}',
                'Host': 'bigdata.lawyee.net',
                'Origin': 'https://bigdata.lawyee.net',
                'Referer': 'https://bigdata.lawyee.net/szyj/search/search',
                'Sec-Fetch-Dest': 'empty',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Site': 'same-origin',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36',
                'X-Requested-With': 'XMLHttpRequest',
                'sec-ch-ua': '"Not(A:Brand";v="99", "Google Chrome";v="133", "Chromium";v="133"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"'
            }

            # 更新后的payload
            payload = {
                "dataSource": "flfg",
                "topicId": "",
                "ids": "",
                "dimSolrTable": "",
                "pageNum": page,
                "pageSize": page_size,
                "isSearch": "false",
                "isChartSearch": "false",
                "dimDataSource": "",
                "isDimSearch": "false",
                "dataQueryCond": "eyJjb25kaXRpb24iOiJBTkQiLCJydWxlcyI6W3siaWQiOjM0LCJmaWVsZCI6MzQsInR5cGUiOiJzdHJpbmciLCJpbnB1dCI6InBvcGN1c3RvbXNlbGVjdHRyZWUiLCJvcGVyYXRvciI6ImluIiwidmFsdWUiOiIwMTAyMDEifV0sInZhbGlkIjp0cnVlfQ=="
            }

            response = self.session.post(
                self.api_url,
                headers=headers,
                data=payload,  # 使用data而不是json，因为Content-Type是form-urlencoded
                timeout=30
            )

            # 检查响应状态，如果是未登录或权限失效，尝试重新登录
            if response.status_code == 401 or response.status_code == 403:
                logging.warning("登录已过期或无权限，尝试重新登录")
                if self.login():
                    # 更新 headers 中的 cookie
                    headers['Cookie'] = f'web={self.auth_info["web_cookie"]}'
                    # 重新发送请求
                    response = self.session.post(
                        self.api_url,
                        headers=headers,
                        data=payload,  # 使用data而不是json，因为Content-Type是form-urlencoded
                        timeout=30
                    )
                else:
                    logging.error("重新登录失败")
                    return None


            if response.status_code == 200:
                try:
                    data = json.loads(response.text)
                    print(data)
                    # data = data.get('data', {})
                    if isinstance(data, dict) and 'data' in data:
                        return data['data']
                    else:
                        logging.error(f"API返回数据格式错误: {data}")
                        return None
                except json.JSONDecodeError as e:
                    logging.error(f"JSON解析失败: {str(e)}")
                    return None
            else:
                logging.error(f"HTTP请求失败: {response.status_code}")
                return None

        except Exception as e:
            logging.error(f"请求异常: {str(e)}")
            return None

    def transform_data(self, original_data: Dict) -> Dict:
        """转换原始数据为规范化的JSON格式"""
        # 初始化结果字典
        result = {
            "ftmc": "",  # 法规标题
            "fgzh": "",  # 法规文号
            "fbjg": "",  # 颁布机构
            "fbrq": "",  # 颁布日期
            "tgrq": "",  # 通过日期
            "sxrq": "",  # 生效日期
            "preface": "",  # 前言
            "presidentOrder": {
                "orderName": "中华人民共和国主席令",
                "orderNo": "",
                "orderContent": "",
                "signatory": "",
                "signDate": ""
            },
            "hasEdition": False,  # 是否有编
            "content": []  # 内容结构
        }

        # 处理基础信息
        for item in original_data.get("result", []):
            try:
                if item.get('type') == 'a' and item.get('pId') == 'n1':
                    name = item.get('name', '').split(' : ')[0] if ' : ' in item.get('name', '') else item.get('name', '')
                    value = item.get('value', '')

                    if name == "法规标题":
                        result["ftmc"] = value
                    elif name == "法规文号":
                        result["fgzh"] = value
                        result["presidentOrder"]["orderNo"] = value.split("第")[-1]
                    elif name == "颁布机构":
                        result["fbjg"] = value
                    elif name == "颁布日期":
                        result["fbrq"] = self.format_date(value)
                        result["presidentOrder"]["signDate"] = self.format_date(value)
                    elif name == "通过日期":
                        result["tgrq"] = self.format_date(value)
                    elif name == "实施日期":
                        result["sxrq"] = self.format_date(value)
            except Exception as e:
                logging.warning(f"处理基础信息时出错: {str(e)}, item: {item}")
                continue

        # 构建前言（从全文分段下寻找题注）
        preface_parts = []
        items = original_data.get("result", [])
        for i, item in enumerate(items):
            try:
                if item.get('isParent') and item.get('name') == "全文分段" and item.get('id'):
                    current_type = None
                    current_content = None
                    
                    # 遍历同一个 pId 下的所有项
                    for sub_item in items:
                        if sub_item.get('pId') == item.get('id'):
                            if '类型' in sub_item.get('name', ''):
                                current_type = sub_item.get('value')
                            elif '内容' in sub_item.get('name', '') and current_type == '题注':
                                preface_parts.append(sub_item.get('value', ''))
                                current_type = None
            except Exception as e:
                logging.warning(f"处理题注时出错: {str(e)}, item: {item}")
                continue
        
        result["preface"] = "\n".join(preface_parts) if preface_parts else ""

        # 构建主席令内容
        result["presidentOrder"]["orderContent"] = (
            f"《{result['ftmc']}》已由{result['fbjg']}于{result['tgrq']}通过，"
            f"现予公布，自{result['sxrq']}起施行。"
        )
        result["presidentOrder"]["signatory"] = "中华人民共和国主席 江泽民"

        # 构建内容结构
        current_edition = None
        current_chapter = None
        current_section = None
        current_article = None
        
        print("开始处理内容...")
        items = original_data.get("result", [])
        for i, item in enumerate(items):
            try:
                if item['type'] == 'a':
                    name = item.get('name', '')
                    content = item.get('value', '')
                    
                    # 检查是否是编、章、节、条的标识
                    if '条文编项' in name:
                        if '编' in content:
                            edition_match = re.match(r'第(\d+)编', content)
                            if edition_match:
                                # 获取标题（在下一个条目中）
                                title = items[i + 1]['value'] if i + 1 < len(items) else ""
                                current_edition = {
                                    "type": "edition",
                                    "code": f"第{self.number_to_chinese(edition_match.group(1))}编",
                                    "title": title,
                                    "children": []
                                }
                                result["content"].append(current_edition)
                                current_chapter = None
                                current_section = None
                        
                        elif '章' in content:
                            chapter_match = re.match(r'第(\d+)章', content)
                            if chapter_match:
                                # 获取标题（在下一个条目中）
                                title = items[i + 1]['value'] if i + 1 < len(items) else ""
                                current_chapter = {
                                    "type": "chapter",
                                    "code": f"第{self.number_to_chinese(chapter_match.group(1))}章",
                                    "title": title,
                                    "children": []
                                }
                                if current_edition:
                                    current_edition["children"].append(current_chapter)
                                else:
                                    result["content"].append(current_chapter)
                                current_section = None
                        
                        elif '节' in content:
                            section_match = re.match(r'第(\d+)节', content)
                            if section_match:
                                # 获取标题（在下一个条目中）
                                title = items[i + 1]['value'] if i + 1 < len(items) else ""
                                current_section = {
                                    "type": "section",
                                    "code": f"第{self.number_to_chinese(section_match.group(1))}节",
                                    "title": title,
                                    "children": []
                                }
                                if current_chapter:
                                    current_chapter["children"].append(current_section)
                        
                        elif '条' in content:
                            article_match = re.match(r'第(\d+)条', content)
                            if article_match:
                                current_article = {
                                    "type": "article",
                                    "code": f"第{self.number_to_chinese(article_match.group(1))}条",
                                    "content": "",
                                    "sxrq": ""
                                }
                                
                                # 获取当前条的 pId
                                current_pid = item.get('pId')
                                
                                # 遍历同一个 pId 下的所有项
                                for sub_item in items:
                                    if sub_item.get('pId') == current_pid:
                                        if '法条内容' in sub_item.get('name', ''):
                                            current_article["content"] = sub_item.get('value', '')
                                        elif '条文生效起始日期' in sub_item.get('name', ''):
                                            current_article["sxrq"] = self.format_date(sub_item.get('value', ''))
                                
                                # 添加到对应的层级
                                if current_section:
                                    current_section["children"].append(current_article)
                                elif current_chapter:
                                    current_chapter["children"].append(current_article)
                                elif current_edition:
                                    current_edition["children"].append(current_article)
                                else:
                                    result["content"].append(current_article)
            except Exception as e:
                logging.warning(f"处理内容时出错: {str(e)}, item: {item}")
                continue

        print(f"处理完成，content长度: {len(result['content'])}")
        return result

    def analyze_structure(self, data: List[Dict]) -> Dict:
        """分析文档结构"""
        structure = {
            "has_edition": False,
            "has_chapter": False,
            "has_section": False
        }
        
        for item in data:
            if item['type'] == 'a' and 'name' in item:
                if '编' in item.get('name', '') and '条' not in item.get('name', ''):
                    structure["has_edition"] = True
                elif '章' in item.get('name', '') and '条' not in item.get('name', ''):
                    structure["has_chapter"] = True
                elif '节' in item.get('name', '') and '条' not in item.get('name', ''):
                    structure["has_section"] = True
        
        return structure

    def number_to_chinese(self, num_str: str) -> str:
        """将数字转换为中文数字"""
        num = int(num_str)
        if num <= 0:
            return num_str
        
        chinese_nums = ['零', '一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
        
        if num < 10:
            return chinese_nums[num]
        elif num < 20:
            if num == 10:
                return chinese_nums[10]
            return chinese_nums[10] + chinese_nums[num % 10]
        elif num < 100:
            if num % 10 == 0:
                return chinese_nums[num // 10] + chinese_nums[10]
            return chinese_nums[num // 10] + chinese_nums[10] + chinese_nums[num % 10]
        
        return num_str  # 对于大于99的数字，保持原样

    def extract_article_code(self, text: str) -> str:
        """提取条文编号并转换为中文"""
        import re
        pattern = r'第(\d+)条'
        match = re.search(pattern, text)
        if match:
            num = match.group(1)
            return f"第{self.number_to_chinese(num)}条"
        return ""

    def format_date(self, date_str: str) -> str:
        """格式化日期 YYYYMMDD -> YYYY-MM-DD"""
        if len(date_str) == 8:
            return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        return date_str

    def belongs_to_bian(self, zhang_number: str, bian_number: str) -> bool:
        """判断章属于哪个编"""
        # 可以根据实际的编号规则来实现
        return True

    def belongs_to_zhang(self, tiao_number: str, zhang_number: str) -> bool:
        """判断条属于哪个章"""
        # 可以根据实际的编号规则来实现
        return True

    def save_transformed_data(self, data: Dict) -> None:
        """保存转换后的数据到文件"""
        try:
            with open(self.result_file, "a", encoding="utf-8") as f:
                json_line = json.dumps(data, ensure_ascii=False)
                f.write(json_line + "\n")
            logging.info(f"成功保存转换后的数据到: {self.result_file}")
        except Exception as e:
            logging.error(f"保存转换后的数据时出错: {str(e)}")

    def batch_fetch(self, start_page=1, end_page=1):
        """批量获取数据"""
        try:
            for page in range(start_page, end_page + 1):
                logging.info(f"正在获取第 {page} 页数据")
                
                # 获取分页列表数据
                page_data = self.fetch_data(page=page, page_size=100)
                print(end_page)
                if  end_page == 1:
                    end_page = page_data["total"]/100+1
                    print("共有" + end_page.__str__() + "页")
                    print("共有" + page_data["total"].__str__() + "条")

                if not page_data:
                    logging.error(f"获取第 {page} 页数据失败")
                    continue

                # 保存到Excel
                if self.save_to_excel(page_data):
                    logging.info(f"第 {page} 页数据已保存到Excel")
                else:
                    logging.error(f"第 {page} 页数据保存到Excel失败")

                # 处理当前页的每条法律数据
                for law_item in page_data.get('list', []):
                    try:
                        row_id = law_item.get('id')
                        if not row_id:
                            continue

                        # 检查法律是否失效
                        law_status = law_item.get('s18_s', '')  # s18_s 字段表示法规效力
                        if law_status == '失效':
                            logging.info(f"法律ID: {row_id} 已失效，跳过获取详细信息")
                            continue

                        # 获取法律详细信息
                        logging.info(f"正在获取法律ID: {row_id} 的详细信息")
                        detail_data = self.fetch_detail(row_id)
                        
                        if detail_data:
                            # 转换数据格式
                            transformed_data = self.transform_data(detail_data)
                            # 保存转换后的数据
                            self.save_transformed_data(transformed_data)
                            logging.info(f"成功处理并保存法律ID: {row_id} 的数据")
                        
                        # 添加延时，避免请求过快
                        time.sleep(5)

                    except Exception as e:
                        logging.error(f"处理法律ID: {row_id} 时出错: {str(e)}")
                        continue

                logging.info(f"第 {page} 页数据处理完成")
                # 页面间添加适当延时
                time.sleep(2)

            logging.info("批量获取数据完成")
            return True

        except Exception as e:
            logging.error(f"批量获取数据时出错: {str(e)}")
            return False

    def save_data(self, data, output_dir='output'):
        """保存数据到文件"""
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.join(output_dir, f'law_data_{timestamp}.json')

        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logging.info(f"数据已保存到: {filename}")
            return True
        except Exception as e:
            logging.error(f"保存数据时出错: {str(e)}")
            return False

    def save_to_excel(self, data, output_file='output/law_data.xlsx'):
        """追加保存数据到Excel文件"""
        try:
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_file), exist_ok=True)

            # 提取新数据
            new_data = []
            for item in data['list']:
                row_data = {
                    '法规标题': item.get('s4_s', ''),
                    '法规层级': item.get('s31_s', ''),
                    '颁布机构': item.get('s7_s', ''),
                    '施行年份': item.get('s16_i', ''),
                    '法规效力': item.get('s18_s', '')
                }
                new_data.append(row_data)

            new_df = pd.DataFrame(new_data)

            # 检查文件是否存在
            if os.path.exists(output_file):
                # 读取现有数据
                existing_df = pd.read_excel(output_file)

                # 合并现有数据和新数据
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)

                # 删除重复数据（基于法规标题）
                combined_df = combined_df.drop_duplicates(subset=['法规标题'], keep='last')
            else:
                combined_df = new_df

            # 创建Excel writer对象
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            combined_df.to_excel(writer, index=False, sheet_name='法规数据')

            # 获取工作表
            worksheet = writer.sheets['法规数据']

            # 设置列宽
            worksheet.column_dimensions['A'].width = 50  # 法规标题
            worksheet.column_dimensions['B'].width = 20  # 法规层级
            worksheet.column_dimensions['C'].width = 30  # 颁布机构
            worksheet.column_dimensions['D'].width = 15  # 施行年份
            worksheet.column_dimensions['E'].width = 15  # 法规效力

            # 设置表头样式
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 设置单元格对齐方式
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # 保存文件
            writer.close()

            logging.info(f"数据已追加到Excel文件: {output_file}")
            logging.info(f"当前共有 {len(combined_df)} 条记录")
            return True

        except Exception as e:
            logging.error(f"保存Excel文件时出错: {str(e)}")
            return False





def main():


    username = "levi"
    password = "Runningwei@6"

    fetcher = LawDataFetcher(username, password)

    try:
        logging.info("开始获取法律数据...")
        data = fetcher.batch_fetch(start_page=1, end_page=8)

        if data:
            if fetcher.save_data(data):
                logging.info("数据获取和保存完成")
            else:
                logging.error("数据保存失败")
        else:
            logging.error("没有获取到数据")

    except Exception as e:
        logging.error(f"程序执行出错: {str(e)}")


if __name__ == "__main__":
    main()
